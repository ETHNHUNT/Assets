import json, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

d = json.load(open("dataset.json"))
COLS = [("record_type","Record Type",14),("name","Name",26),("prompt_text","Prompt Text",90),
        ("description","Description",48),("model_or_effect","Model / Motion Effect",22),
        ("tool_type","Tool Type",24),("generation_style","Generation Style",30),
        ("visual_subject","Visual Subject",30),("category","Category",18),
        ("preset_name","Preset",18),("aspect_ratio","Aspect Ratio",12),
        ("duration_sec","Duration (s)",12),("quality","Quality",10),("badges","Badges",22),
        ("word_count","Words",8),("char_count","Chars",8),("confidence","Confidence",12),
        ("asset_count","Assets",8),("asset_type","Asset Type",12),
        ("thumb_path","Thumbnail (repo path)",34),("full_res_url","Full-Res Asset URL",52),
        ("poster_url","Poster URL",46),("media_pairing","Asset Pairing",14),
        ("recreate_model","Recreate Model",20),("lesson_title","Lesson",30),
        ("timestamp_in_lesson","Lesson Ts (s)",12),
        ("site_section","Site Section",18),("extraction_source","Extraction Source",18),
        ("media_url","Sample Media URL",42),("source_url","Source Page URL",52)]

HDR_FILL = PatternFill("solid", fgColor="1F2937")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

def sheet(ws, rows, name):
    ws.title = name
    ws.append([h for _, h, _ in COLS])
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
    for r in rows:
        ws.append([r.get(k) for k, _, _ in COLS])
    for i, (k, h, wdt) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 26
    n = len(rows) + 1
    for row in ws.iter_rows(min_row=2, max_row=n, max_col=len(COLS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column in (3, 4, 7, 8)))
            cell.border = BORDER
    if n > 1:
        t = Table(displayName=name.replace(" ", "").replace("/", "").replace("-", "")[:28] + "Tbl",
                  ref=f"A1:{get_column_letter(len(COLS))}{n}")
        t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        try: ws.add_table(t)
        except Exception: pass

sheet(wb.active, d, "All Records")
sheet(wb.create_sheet(), [r for r in d if r["record_type"] == "Prompt"], "Prompts")
sheet(wb.create_sheet(), [r for r in d if r["record_type"] == "Preset  Effect"] or
      [r for r in d if r["record_type"] == "Preset / Effect"], "Presets and Effects")

# summary
ws = wb.create_sheet("Summary")
ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 12
row = 1
ws.cell(row=1, column=1, value="Higgsfield.ai Prompt Extraction — Summary").font = Font(bold=True, size=14)
row = 3
def block(title, key, split=False):
    global row
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=11)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E5E7EB")
    ws.cell(row=row, column=2, value="Count").font = Font(bold=True, size=11)
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="E5E7EB")
    row += 1
    c = collections.Counter()
    for r in d:
        v = r.get(key)
        if not v: c["(unspecified)"] += 1; continue
        if split:
            for p in str(v).split("; "): c[p] += 1
        else: c[str(v)] += 1
    for k, v in c.most_common():
        ws.cell(row=row, column=1, value=k); ws.cell(row=row, column=2, value=v); row += 1
    row += 1

ws.cell(row=row, column=1, value="Total records"); ws.cell(row=row, column=2, value=len(d)); row += 2
block("BY TOOL TYPE", "tool_type")
block("BY MODEL / MOTION EFFECT", "model_or_effect")
block("BY GENERATION STYLE", "generation_style", True)
block("BY VISUAL SUBJECT", "visual_subject", True)
block("BY SITE SECTION", "site_section")
block("BY EXTRACTION SOURCE", "extraction_source")
block("BY CONFIDENCE", "confidence")

wb.save("deliverables/higgsfield_prompt_dataset.xlsx")
print("xlsx written")
