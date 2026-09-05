"""End-to-end checks for the Higgsfield dataset build.

Runs from the repository root, against either of two layouts:

  * mid-build  — `dataset.json` and `deliverables/` in the working directory,
                 as `clean.py` and the `build_*.py` scripts leave them;
  * a clone    — neither of those exists, so the committed copies in `data/`
                 are checked instead.

Checks whose inputs are genuinely absent report SKIP. They must never report
PASS: a check that ran over nothing has established nothing.
"""
import json, os, csv, re, sys, glob, collections
sys.path.insert(0, '.')

FAIL = []
def check(name, cond, detail=""):
    print(("  PASS  " if cond else "  FAIL  ") + name + (f" — {detail}" if detail else ""))
    if not cond:
        FAIL.append(name)

def skip(name, reason):
    print(f"  SKIP  {name} — {reason}")

def pick(*paths):
    """First path that exists, or None."""
    return next((p for p in paths if os.path.exists(p)), None)

DATASET = pick("dataset.json", "data/higgsfield_prompt_dataset.json")
if not DATASET:
    sys.exit("no dataset: expected ./dataset.json (a build intermediate) or "
             "data/higgsfield_prompt_dataset.json (committed)")
# The build writes into deliverables/; the committed copies live in data/. Prefer a
# fresh build when one is present, so a rebuild is checked in preference to the
# committed artefacts it is meant to replace.
DELIV = ("deliverables" if os.path.exists("deliverables/higgsfield_prompts_full.csv")
         else "data" if os.path.exists("data/higgsfield_prompts_full.csv") else None)
PAGES = sorted(glob.glob("pages/*.html"))

print(f"dataset:      {DATASET}")
print(f"deliverables: {DELIV or '(none found)'}")
print(f"pages/:       {len(PAGES)} html files\n")

print("1. Coverage / gap closure")
d = json.load(open(DATASET))
check("records >= 1600", len(d) >= 1600, f"{len(d)} records")
if not PAGES:
    # pages/ is a crawl intermediate and is not committed (README section 8), so on a
    # clone there is nothing to re-extract and compare against.
    skip("no uncaptured academy prompts", "pages/ absent — crawl intermediate, not committed")
else:
    from recreate import srcurl
    have = {r["source_url"] for r in d}
    pages = []
    for f in PAGES:
        u = srcurl(open(f, encoding="utf-8", errors="replace").read(3000))
        if u:
            pages.append(u)
    zero = [u for u in pages if u not in have]
    acad_zero = len([u for u in zero if "/academy/" in u])
    # A lesson page with no prompt in it is not a miss -- most are video-only lessons.
    # What matters is whether any prompt on those pages went uncaptured.
    import unicodedata
    from flat_prompts import extract_flat
    def _n(t):
        t = unicodedata.normalize("NFKC", t or "")
        return re.sub(r"[^a-z0-9 ]+", "", re.sub(r"\s+", " ", t).strip().lower())[:400]
    captured = {_n(r["prompt_text"]) for r in d if r.get("prompt_text")}
    uncaptured = 0
    for f in PAGES:
        h = open(f, encoding="utf-8", errors="replace").read()
        u = srcurl(h)
        if not (u and "/academy/" in u and u not in have):
            continue
        for x in extract_flat(h, u):
            if _n(x["prompt"]) not in captured:
                uncaptured += 1
    check("no uncaptured academy prompts", uncaptured == 0,
          f"{uncaptured} uncaptured across {acad_zero} prompt-free lesson pages (was 182)")
flat = len([r for r in d if r["extraction_source"] == "flat_payload"])
check("flat_payload records >= 450", flat >= 450, f"{flat}")

print("\n2. Prompt-to-asset pairing")
withasset = [r for r in d if r.get("full_res_url")]
pct = 100 * len(withasset) / len(d)
check("assets on >= 95% of records", pct >= 95, f"{pct:.1f}%")
pres = [r for r in d if r["record_type"] == "Preset / Effect"]
presa = [r for r in pres if r.get("full_res_url")]
check("presets all have assets", len(presa) == len(pres), f"{len(presa)}/{len(pres)}")

print("\n3. Thumbnails")
man = json.load(open("assets/manifest.json"))
withthumb = [m for m in man if m.get("thumb_path")]
missing = [m for m in withthumb if not os.path.exists(os.path.join("assets", m["thumb_path"]))]
check("no manifest thumb points at a missing file", not missing, f"{len(missing)} missing")
# The converse: entries carrying no thumbnail at all. Every one of these is an asset
# the CDN no longer serves — see tools/recheck_thumbs.py, which re-probes them.
nothumb = [m for m in man if not m.get("thumb_path")]
check("un-thumbed assets stay within the known set", len(nothumb) <= 19,
      f"{len(nothumb)} of {len(man)} assets have no thumbnail (19 known dead on the CDN)")
try:
    from PIL import Image
    bad = []
    for f in glob.glob("assets/thumbs/*.webp"):
        try:
            Image.open(f).verify()
        except Exception:
            bad.append(f)
    check("all thumbnails decode", not bad, f"{len(bad)} corrupt of {len(glob.glob('assets/thumbs/*.webp'))}")
except ImportError:
    skip("all thumbnails decode", "Pillow unavailable — pip install -r requirements.txt")
size = sum(os.path.getsize(f) for f in glob.glob("assets/thumbs/*.webp"))
check("thumbnail set <= 150 MB", size <= 150 * 1048576, f"{size/1048576:.0f} MB")

print("\n4. Deliverables")
if not DELIV:
    skip("deliverables", "neither deliverables/ nor data/ holds higgsfield_prompts_full.csv")
else:
    rows = list(csv.DictReader(open(f"{DELIV}/higgsfield_prompts_full.csv", encoding="utf-8-sig")))
    check("CSV row count matches dataset", len(rows) == len(d), f"{len(rows)} vs {len(d)}")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(f"{DELIV}/higgsfield_prompt_dataset.xlsx")
        check("XLSX has 4 sheets", len(wb.sheetnames) == 4, str(wb.sheetnames))
        check("XLSX row count matches", wb["All Records"].max_row == len(d) + 1,
              f"{wb['All Records'].max_row - 1}")
    except ImportError:
        skip("XLSX checks", "openpyxl unavailable — pip install -r requirements.txt")
    except Exception as e:
        check("XLSX opens", False, str(e)[:60])
    pdf = open(f"{DELIV}/higgsfield_prompt_dataset.pdf", "rb").read()
    check("PDF header/EOF valid", pdf[:8] == b"%PDF-1.4" and pdf.rstrip().endswith(b"%%EOF"))
    cnt = re.findall(rb"/Count\s+(\d+)", pdf)
    check("PDF has pages", bool(cnt) and int(cnt[0]) > 50, f"{cnt[0].decode() if cnt else '0'} pages")

print("\n5. Gallery")
g = open("assets/gallery.html", encoding="utf-8").read()
srcs = re.findall(r'<img[^>]*src="([^"]+)"', g)
gm = [s for s in srcs if not os.path.exists(os.path.join("assets", s))]
check("every gallery image exists on disk", not gm, f"{len(gm)} missing of {len(srcs)}")
check("gallery has cards", g.count('<article class="card"') > 100,
      f"{g.count('<article class=' + chr(34) + 'card' + chr(34))} cards")

print("\n" + ("ALL CHECKS PASSED" if not FAIL else f"{len(FAIL)} CHECK(S) FAILED: {FAIL}"))
sys.exit(1 if FAIL else 0)
